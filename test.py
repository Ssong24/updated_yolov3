import argparse
import json

from torch.utils.data import DataLoader

from models import *
from utils.datasets import *
from utils.utils import *
# from collections import defaultdict
import numpy as np
import openpyxl

def test(cfg,
         data,
         weights=None,
         batch_size=16,
         imgsz=416,
         conf_thres=0.001,
         iou_thres=0.6,  # for nms
         save_json=False,
         single_cls=False,
         augment=False,
         model=None,
         dataloader=None,
         multi_label=True,
         data_format='cityscape',
         n_classes=8,
         gt_json='kitti_original_gt.json',
         save_result =False,
         projection='',
         autoanchor=False):

    # Initialize/load model and set device
    if model is None:
        is_training = False
        device, _ = torch_utils.select_device(opt.device, batch_size=batch_size)
        verbose = opt.task == 'test'

        # Remove previous
        for f in glob.glob('test_batch*.jpg'):
            os.remove(f)

        # Initialize model
        model = Darknet(cfg, imgsz)

        # anchor check
        if autoanchor is True:
            num_anchors = 9
            data_dict = parse_data_cfg(data)
            train_path = data_dict['train']
            new_anchors = kmean_anchors(train_path, n=num_anchors, img_size=(imgsz, imgsz), thr=0.35, gen=1000,
                                        data_format=data_format, n_classes=n_classes)
            yolo_index = 0
            for module in model.module_list:
                if isinstance(module, YOLOLayer):
                    module.anchors = torch.Tensor(new_anchors[6 - yolo_index * 3: 9 - yolo_index * 3])
                    yolo_index += 1

        # Load weights
        attempt_download(weights)
        if weights.endswith('.pt'):  # pytorch format
            model.load_state_dict(torch.load(weights, map_location=device)['model'], strict=False)
        else:  # darknet format
            load_darknet_weights(model, weights)

        # Fuse
        model.fuse()
        model.to(device)

        if device.type != 'cpu' and torch.cuda.device_count() > 1:
            model = nn.DataParallel(model)

    else:  # called by train.py
        is_training = True
        device = next(model.parameters()).device  # get model device
        verbose = False

    # Configure run
    data = parse_data_cfg(data)
    nc = 1 if single_cls else int(data['classes'])  # number of classes

    path = data['valid']  # path to test images -- train.py mode
    names = load_classes(data['names'])  # class names
    iouv = torch.linspace(0.5, 0.95, 10).to(device)  # iou vector for mAP@0.5:0.95
    iouv = iouv[0].view(1)  # comment for mAP@0.5:0.95
    niou = iouv.numel()  # 10?

    # Dataloader -- test mode
    if dataloader is None:
        path = data['test']
        dataset = LoadImagesAndLabels(path, imgsz, batch_size, rect=True, single_cls=opt.single_cls, pad=0.5
                                      , data_format=data_format, n_classes=n_classes)
        batch_size = min(batch_size, len(dataset))
        dataloader = DataLoader(dataset,
                                batch_size=batch_size,
                                num_workers=min([os.cpu_count(), batch_size if batch_size > 1 else 0, 8]),
                                pin_memory=True,
                                collate_fn=dataset.collate_fn)

    seen = 0
    model.eval()
    _ = model(torch.zeros((1, 3, imgsz, imgsz), device=device)) if device.type != 'cpu' else None  # run once
    coco91class = coco80_to_coco91_class()
    s = ('%20s' + '%10s' * 6) % ('Class', 'Images', 'Targets', 'P', 'R', 'mAP@0.5', 'F1')
    p, r, f1, mp, mr, map, mf1, t0, t1 = 0., 0., 0., 0., 0., 0., 0., 0., 0.
    loss = torch.zeros(4, device=device)
    jdict, stats, ap, ap_class = [], [], [], []

    for batch_i, (imgs, targets, paths, shapes) in enumerate(tqdm(dataloader, desc=s)):
        imgs = imgs.to(device).float() / 255.0  # uint8 to float32, 0 - 255 to 0.0 - 1.0
        targets = targets.to(device)  # img id, class id, x, y, w, h
        nb, _, height, width = imgs.shape  # batch size, channels, height, width
        whwh = torch.Tensor([width, height, width, height]).to(device)

        # Disable gradients
        with torch.no_grad():
            # Run model
            t = torch_utils.time_synchronized()
            inf_out, train_out = model(imgs, augment=augment)  # inference and training outputs
            t0 += torch_utils.time_synchronized() - t

            # Compute loss
            if is_training:  # if model has loss hyperparameters
                loss += compute_loss(train_out, targets, model, loss_size=False)[1][:4]  # GIoU, obj, cls, size

            # Run NMS
            t = torch_utils.time_synchronized()
            output = non_max_suppression(inf_out, conf_thres=conf_thres, iou_thres=iou_thres, multi_label=multi_label)
            t1 += torch_utils.time_synchronized() - t

        # Statistics per image
        for si, pred in enumerate(output):  # pred: (x1,y1,x2,y2,conf,cls)
            labels = targets[targets[:, 0] == si, 1:]  # targets[:,0] - image id

            nl = len(labels)
            tcls = labels[:, 0].tolist() if nl else []  # target class
            seen += 1

            if pred is None:
                if nl:
                    stats.append((torch.zeros(0, niou, dtype=torch.bool), torch.Tensor(), torch.Tensor(), tcls))
                continue

            # Clip boxes to image bounds
            clip_coords(pred, (height, width))

            # Append to pycocotools JSON dictionary
            if save_json:
                # [{"image_id": 42, "category_id": 18, "bbox": [258.15, 41.29, 348.26, 243.78], "score": 0.236}, ...
                if data_format == 'kitti':
                    image_id = int(Path(paths[si]).stem.split('_')[-1])
                elif data_format == 'cityscape':
                    image_id = get_cityscape_img_id(paths[si])
                elif data_format == 'fisheye':
                    image_id = int(Path(paths[si]).stem.split('-')[1] + Path(paths[si]).stem.split('-')[2])
                elif data_format == 'woodscape':
                    image_id = int(Path(paths[si]).stem.split('_')[0])

                box = pred[:, :4].clone()  # xyxy

                scale_coords(imgs[si].shape[1:], box, shapes[si][0], shapes[si][1])  # to original shape
                box = xyxy2xywh(box)  # xywh
                box[:, :2] -= box[:, 2:] / 2  # xy center to top-left corner
                for p, b in zip(pred.tolist(), box.tolist()):
                    jdict.append({'image_id': image_id,
                                  'category_id': coco91class[int(p[5])],
                                  'bbox': [round(x, 3) for x in b],
                                  'score': round(p[4], 5)})

            # Assign all predictions as incorrect
            correct = torch.zeros(pred.shape[0], niou, dtype=torch.bool, device=device)

            if nl:
                detected = []  # target indices
                tcls_tensor = labels[:, 0]

                # target boxes
                tbox = xywh2xyxy(labels[:, 1:5]) * whwh

                # Per target class
                for cls in torch.unique(tcls_tensor):
                    ti = (cls == tcls_tensor).nonzero().view(-1)  # target indices
                    pi = (cls == pred[:, 5]).nonzero().view(-1)  # prediction indices

                    # Search for detections
                    if pi.shape[0]:
                        # Prediction to target ious
                        ious, i = box_iou(pred[pi, :4], tbox[ti]).max(1)  # best ious, indices

                        # Append detections
                        for j in (ious > iouv[0]).nonzero():
                            d = ti[i[j]]  # detected target
                            if d not in detected:
                                detected.append(d)
                                correct[pi[j]] = ious[j] > iouv  # iou_thres is 1xn
                                if len(detected) == nl:  # all targets already located in image
                                    break

            # Append statistics (correct, conf, pcls, tcls)
            stats.append((correct.cpu(), pred[:, 4].cpu(), pred[:, 5].cpu(), tcls))

        # Plot images
        # if batch_i < 1:
        #     f = 'test_batch%g_gt.jpg' % batch_i  # filename
        #     plot_images(imgs, targets, paths=paths, names=names, fname=f)  # ground truth
        #     f = 'test_batch%g_pred.jpg' % batch_i
        #     plot_images(imgs, output_to_target(output, width, height), paths=paths, names=names, fname=f)  # predictions

    # Compute statistics
    stats = [np.concatenate(x, 0) for x in zip(*stats)]  # to numpy
    if len(stats):
        p, r, ap, f1, ap_class = ap_per_class(*stats)
        if niou > 1:  # What is this?
            p, r, ap, f1 = p[:, 0], r[:, 0], ap.mean(1), ap[:, 0]  # [P, R, AP@0.5:0.95, AP@0.5]
        print('ap: ', ap)
        print('ap[:, 0]', ap[:,0])
        mp, mr, map, mf1 = p.mean(), r.mean(), ap.mean(), f1.mean()

        nt = np.bincount(stats[3].astype(np.int64), minlength=nc)  # number of targets per class
    else:
        nt = torch.zeros(1)

    # Print results
    pf = '%20s' + '%10.3g' * 6  # print format
    print(pf % ('all', seen, nt.sum(), mp, mr, map, mf1))  # result of 'all'

    # Print results per class
    if verbose and nc > 1 and len(stats):
        for i, c in enumerate(ap_class):
            print(pf % (names[c], seen, nt[c], p[i], r[i], ap[i], f1[i]))

    # Print speeds
    if verbose or save_json:
        t = tuple(x / seen * 1E3 for x in (t0, t1, t0 + t1)) + (imgsz, imgsz, batch_size)  # tuple
        print('Speed: %.1f/%.1f/%.1f ms inference/NMS/total per %gx%g image at batch-size %g' % t)

    # Save JSON
    if save_json and len(jdict) and map:
        print('\nCOCO mAP with pycocotools...')
        if data_format == 'kitti':
            imgIds = [int(Path(x).stem.split('_')[-1]) for x in dataloader.dataset.img_files]
        elif data_format == "cityscape":
            imgIds = [get_cityscape_img_id(x) for x in dataloader.dataset.img_files]
        elif data_format == "fisheye":
            imgIds = [int(Path(x).stem.split('-')[1] + Path(x).stem.split('-')[2].split('_')[0]) for x in dataloader.dataset.img_files]
        elif data_format == "woodscape":
            imgIds = [int(Path(x).stem.split('_')[0]) for x in dataloader.dataset.img_files]
        else:
            print('Please set the dataset')
            exit(-1)
        
        json_folder = 'results/' + data_format + '/' + projection + weights.split('\\')[-3]
        json_file = json_folder + '/results.json'

        print('weights: ', weights)
        print('json_file: ', json_folder + '/results.json')
        print('gt_json: ', gt_json)

        with open(json_file, 'w') as file:
            json.dump(jdict, file)

        from pycocotools.coco import COCO
        from pycocotools.cocoeval import COCOeval

        # https://github.com/cocodataset/cocoapi/blob/master/PythonAPI/pycocoEvalDemo.ipynb
        cocoGt = COCO(gt_json)  # initialize COCO ground truth api
        cocoDt = cocoGt.loadRes(json_file)  # initialize COCO pred api

        cocoEval = COCOeval(cocoGt, cocoDt, 'bbox')
        cocoEval.params.imgIds = imgIds  # [:32]  # only evaluate these images

        cocoEval.evaluate()
        cocoEval.accumulate()
        cocoEval.summarize()

        if save_result:
            dic_col_cocoapi = {
                # COCOAPI
                'Model': 1, 'AP': 2, 'AP50': 3, 'AP75': 4,
                'AP_s-small': 5, 'AP_v-small': 6, 'AP_small': 7, 'AP_medium': 8, 'AP_large': 9,
                'AR-1': 10, 'AR-10': 11, 'AR-100': 12,
                'AR_s-small': 13, 'AR_v-small': 14, 'AR_small': 15, 'AR_medium': 16, 'AR_large': 17
            }
            dic_col_kitti = {
                # Class-
                'Model': 1, 'all': 2,
                'Pedestrian': 3, 'Person_sitting': 4, 'Cyclist': 5, 'Car': 6, 'Van': 7, 'Truck': 8, 'Tram': 9
            }
            dic_col_cityscape = {
                'Model': 1, 'all': 2,
                'person': 3, 'rider': 4, 'car': 5, 'truck': 6, 'bus': 7, 'train': 8, 'motorcycle': 9, 'bicycle': 10,
                'caravan': 11, 'trailer': 12
            }
            dic_col_fisheye = {
                'Model': 1, 'all': 2,
                'pedestrian': 3, 'rider': 4, 'person_sitting': 5, 'bicycle': 6, 'motorcycle': 7, 'car': 8, 'van': 9,
                'truck': 10, 'bus': 11
            }

            dic_col_woodscape = {
                'Model': 1, 'all': 2,
                'vehicle': 3, 'person': 4, 'rider': 5, 'bicycle': 6, 'motorcycle': 7, 'traffic_light': 8, 'traffic_sign': 9
            }

            if data_format =='cityscape':
                excel_name = 'results/cs_val_result.xlsx'
                dic_col_dataset = dic_col_cityscape
            elif data_format == 'kitti':
                excel_name = 'results/kitti_test_result.xlsx'
                dic_col_dataset = dic_col_kitti
            elif data_format == 'fisheye':
                excel_name = 'results/fisheye_test_result.xlsx'
                dic_col_dataset = dic_col_fisheye
            elif data_format == "woodscape":
                excel_name = 'results/woodscape_test_result.xlsx'
                dic_col_dataset = dic_col_woodscape

            f_excel = openpyxl.load_workbook(excel_name)

            # Automatically find column number
            # for wsheet in f_excel:
            #     for c in range(1, 20):
            #         if wsheet.cell(row=1, column=c).value == 'Model':
            #             break  # dir_col_num['Model'] = c

            sheet_cocoapi = f_excel["cocoapi"]
            sheet_row = 1

            model_version_name = weights.replace("\\", '/').split('/')[-3]  # format: results/[dir name]/weights/best.pt

            while True:
                if sheet_cocoapi.cell(row=sheet_row, column=1).value == None:
                    start_row = sheet_row
                    break
                sheet_row += 1

            # Write model_version_name in each worksheet except "Targets"
            for wsheet in f_excel:
                if wsheet != f_excel["Targets"]:
                    wsheet.cell(row=start_row, column=dic_col_cocoapi['Model']).value = model_version_name

            # write result value in each specific worksheet
            write_result_in_excel(f_excel, "class-AP50", start_row, dic_col_dataset, True, names, map, ap)
            write_result_in_excel(f_excel, "class-Precision", start_row, dic_col_dataset, True, names, mp, p)
            write_result_in_excel(f_excel, "class-Recall", start_row, dic_col_dataset, True, names, mr, r)
            write_result_in_excel(f_excel, 'cocoapi', start_row, dic_col_cocoapi, is_class_result=False, coco_values=cocoEval.stats)
            write_result_in_excel(f_excel, 'Targets', start_row, dic_col_dataset, True, names, all_value=nt.sum(), class_values=np.reshape(nt, (nc, -1)))

            # Update the value of that file excel
            f_excel.save(excel_name)
            f_excel.close()

    # Return results
    maps = np.zeros(nc) + map
    for i, c in enumerate(ap_class):
        maps[c] = ap[i]
    return (mp, mr, map, mf1, *(loss.cpu() / len(dataloader)).tolist()), maps, names


def write_result_in_excel(file_excel, worksheet_name, start_row, dir_col,
                          is_class_result=False, class_names=None, all_value=None, class_values=None, coco_values=None):
    if is_class_result:
        for wsheet in file_excel:
            if wsheet == file_excel[worksheet_name]:
                wsheet.cell(row=start_row, column=dir_col['all']).value = all_value
                for i, c in enumerate(class_names):
                     wsheet.cell(row=start_row, column=dir_col[c]).value = class_values[i][0]
    else:
        for i, stat in enumerate(coco_values):
            file_excel[worksheet_name].cell(row=start_row, column=i+2).value = stat


if __name__ == '__main__':
    parser = argparse.ArgumentParser(prog='test.py')
    parser.add_argument('--cfg', type=str, default='cfg/yolov3-spp.cfg', help='*.cfg path')
    parser.add_argument('--data', type=str, default='data/coco2014.data', help='*.data path')
    parser.add_argument('--weights', type=str, default='input/pretrained_weights/darknet53.conv.74', help='weights path')
    parser.add_argument('--batch-size', type=int, default=20, help='size of each image batch')
    parser.add_argument('--img-size', type=int, default=512, help='inference size (pixels)')
    parser.add_argument('--conf-thres', type=float, default=0.001, help='object confidence threshold')
    parser.add_argument('--iou-thres', type=float, default=0.6, help='IOU threshold for NMS')
    parser.add_argument('--save-json', action='store_true', help='save a cocoapi-compatible JSON results file')
    parser.add_argument('--task', default='test', help="'test', 'study', 'benchmark'")
    parser.add_argument('--device', default='0', help='device id (i.e. 0 or 0,1) or cpu')
    parser.add_argument('--single-cls', action='store_true', help='train as single-class dataset')
    parser.add_argument('--augment', action='store_true', help='augmented inference')
    # Song's argument
    parser.add_argument('--data-format', type=str, default='cityscape', help='etri | kitti | cityscape | coco')
    parser.add_argument('--n-classes', type=int, default=8, help='number of classes')
    parser.add_argument('--gt-json', type=str, default='kitti_original.json', help='Name of grundtruth label with coco format ')
    parser.add_argument('--save-result', action='store_true', help='save test result in excel file.')
    parser.add_argument('--projection', type=str, default='', help='fe/ | equi/ | songCylin/')
    parser.add_argument('--autoanchor', action='store_true', help='Enable autoanchor check')
    opt = parser.parse_args()
    opt.save_json = opt.save_json or any([x in opt.data for x in ['coco.data', 'coco2014.data', 'coco2017.data']])
    opt.cfg = check_file(opt.cfg)  # check file
    opt.data = check_file(opt.data)  # check file

    print(opt)

    # task = 'test', 'study', 'benchmark'
    if opt.task == 'test':  # (default) test normally
        test(opt.cfg,
             opt.data,
             opt.weights,
             opt.batch_size,
             opt.img_size,
             opt.conf_thres,
             opt.iou_thres,
             opt.save_json,
             opt.single_cls,
             opt.augment,
             data_format=opt.data_format,
             n_classes=opt.n_classes,
             gt_json=opt.gt_json,
             save_result=opt.save_result,
             projection=opt.projection,
             autoanchor=opt.autoanchor)

    elif opt.task == 'benchmark':  # mAPs at 256-640 at conf 0.5 and 0.7
        y = []
        for i in list(range(256, 640, 128)):  # img-size
            for j in [0.6, 0.7]:  # iou-thres
                t = time.time()
                r = test(opt.cfg, opt.data, opt.weights, opt.batch_size, i, opt.conf_thres, j, opt.save_json)[0]
                y.append(r + (time.time() - t,))
        np.savetxt('benchmark.txt', y, fmt='%10.4g')  # y = np.loadtxt('study.txt')
